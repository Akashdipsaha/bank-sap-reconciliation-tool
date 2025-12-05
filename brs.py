import streamlit as st
import os, warnings
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ----------------------------------------------------
# 1. Streamlit Page Setup
# ----------------------------------------------------
st.set_page_config(
    page_title="Bank Reconciliation System", 
    page_icon="🏦", 
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ----------------------------------------------------
# 2. Session State Management
# ----------------------------------------------------
if 'processed_data' not in st.session_state:
    st.session_state.processed_data = None
if 'excel_buffer' not in st.session_state:
    st.session_state.excel_buffer = None
if 'metrics' not in st.session_state:
    st.session_state.metrics = {}

def reset_app():
    """Clears the session state to allow new uploads"""
    st.session_state.processed_data = None
    st.session_state.excel_buffer = None
    st.session_state.metrics = {}

# ----------------------------------------------------
# 3. 🌊 CSS: Styling & Animation
# ----------------------------------------------------
st.markdown("""
<style>
/* General App Background */
html, body, .stApp {
    background-color: #f8fafc;
    font-family: "Segoe UI", sans-serif;
    color: #0f172a;
}

/* --- VISIBILITY FIX FOR DARK MODE (LIGHT BLUE TEXT) --- */
/* Target File Uploader Text (Browse files, Drag and drop, Limits) */
[data-testid="stFileUploaderDropzone"] div,
[data-testid="stFileUploaderDropzone"] span,
[data-testid="stFileUploaderDropzone"] small,
[data-testid="stFileUploaderDropzone"] button {
    color: #0ea5e9 !important; /* Light Blue */
}

/* Target Labels (e.g., 'Select Sheet') */
.stSelectbox label, 
.stSelectbox div[data-testid="stMarkdownContainer"] p,
.stFileUploader label,
div[data-testid="stMarkdownContainer"] p {
    color: #0ea5e9 !important;
}

/* Ensure the upload button inside the dropzone is visible */
section[data-testid="stFileUploader"] button {
    color: #0ea5e9 !important;
    border-color: #0ea5e9 !important;
}
/* ------------------------------------------------------ */

/* Keyframes */
@keyframes shimmer {
    0% {background-position: 200% center;}
    100% {background-position: -200% center;}
}

/* Header - VERY SLOW animation (60s) */
.app-header {
    background: linear-gradient(120deg, #0f172a 30%, #0ea5e9 50%, #0f172a 70%);
    background-size: 200% auto;
    animation: shimmer 60s linear infinite;
    padding: 15px;
    border-radius: 12px;
    text-align: center;
    box-shadow: 0 10px 25px rgba(14, 165, 233, 0.2);
    margin-bottom: 25px;
}
.app-header h2 { color: #ffffff; margin: 0; font-weight: 700; letter-spacing: 1px; font-size: 24px; }
.app-header p { color: #e0f2fe; margin-top: 5px; font-size: 14px; }

/* Card Styling */
.upload-card {
    background-color: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 12px;
    padding: 15px;
    margin-bottom: 15px;
    box-shadow: 0 4px 6px rgba(0,0,0,0.02);
    transition: all 0.3s ease;
}
.upload-card:hover {
    border-color: #38bdf8;
    box-shadow: 0 0 15px rgba(14, 165, 233, 0.1);
}
.card-header {
    font-size: 1rem;
    font-weight: 700;
    color: #334155;
    margin-bottom: 10px;
    border-bottom: 2px solid #f1f5f9;
    padding-bottom: 5px;
}

/* Buttons (Regular & Download) */
div.stButton > button, .stDownloadButton > button {
    background: linear-gradient(90deg, #0284c7 0%, #38bdf8 50%, #0284c7 100%);
    background-size: 200% auto;
    animation: shimmer 30s linear infinite;
    color: #ffffff !important; /* White Text Forced */
    font-weight: 600; border: none;
    border-radius: 8px; padding: 10px 20px;
    font-size: 14px; width: 100%;
    box-shadow: 0 4px 10px rgba(2, 132, 199, 0.3);
    transition: transform 0.2s;
}

/* Force Text inside Buttons (p tags) to be White */
div.stButton > button p, .stDownloadButton > button p {
    color: #ffffff !important;
}

div.stButton > button:hover, .stDownloadButton > button:hover {
    transform: translateY(-2px);
    box-shadow: 0 6px 15px rgba(56, 189, 248, 0.5);
    color: #ffffff !important;
}

/* Reset Button */
div.stButton > button.reset-btn {
    background: linear-gradient(90deg, #ef4444 0%, #f87171 50%, #ef4444 100%);
}

/* Results Styling */
[data-testid="stMetricValue"] { color: #0284c7 !important; font-weight: 700; }
.stDataFrame { border-radius: 10px; overflow: hidden; border: 1px solid #e2e8f0; }
.legend-box {
    background: #f1f5f9;
    padding: 10px 15px; border-radius: 8px; margin-top: 15px;
    color: #475569; font-size: 13px; border-left: 4px solid #0ea5e9;
}
.empty-state {
    text-align: center; padding: 40px; color: #94a3b8;
    background: white; border: 2px dashed #cbd5e1; border-radius: 12px;
}
</style>
""", unsafe_allow_html=True)

# ----------------------------------------------------
# 4. Header & Sidebar
# ----------------------------------------------------
st.markdown("""
<div class='app-header'>
  <h2>🏦 Automated Bank Reconciliation System</h2>
  <p>Secure, Reliable & One-Stop Solution for Bank-to-SAP Ledger Matching</p>
</div>
""", unsafe_allow_html=True)

st.sidebar.header("⚙️ Configuration")
acct_type = st.sidebar.selectbox("Account Type", ["G/L Account", "BRS Account"], index=1)
st.sidebar.markdown("---")

# ----------------------------------------------------
# 5. Processor Class (UPDATED FOR KZ->ZR & HISTORY)
# ----------------------------------------------------
class Processor:
    def __init__(self, b, s, h, a):
        self.b, self.s, self.h, self.a = b, s, h, a
        self.bank_date_col = None
        self.sap_date_col = None
        self.bank_ref_col = None
        self.sap_ref_col = None
        self.sap_type_col = None
        self.orig_bank_ref_col = None 
        self.orig_sap_ref_col = None
        self.df_history = None

    def clean_currency(self, series):
        if series.dtype == 'object':
            return pd.to_numeric(series.astype(str).str.replace(',', '').str.strip(), errors='coerce')
        return pd.to_numeric(series, errors='coerce')

    def clean_ref(self, series):
        if series is None: return None
        return series.astype(str).str.replace(r'^[0]+', '', regex=True).str.strip().str.lower()

    def _prep(self, df):
        if df.empty: return pd.DataFrame()
        header_idx = None
        scan_limit = min(25, len(df))
        
        for i in range(scan_limit):
            row_vals = df.iloc[i].astype(str).str.lower()
            has_date = any(x in row_vals.values for x in ["date", "txn date", "transaction date", "val. date", "posting date", "tran date"])
            has_amt = any(x in row_vals.values for x in ["amount", "debit", "credit", "withdrawal", "deposit", "withdrawals", "deposits", "amount(inr)", "dr amount", "cr amount", "dr/cr", "dr", "cr"])
            
            if has_date and has_amt:
                header_idx = i
                break
        
        if header_idx is None: 
            for i in range(scan_limit):
                row_vals = df.iloc[i].astype(str).str.lower()
                if any(x in row_vals.values for x in ["date", "txn date", "transaction date", "tran date"]):
                    header_idx = i; break

        if header_idx is None:
            try:
                first_cell = pd.to_datetime(df.iloc[0, 0], dayfirst=True, errors='coerce')
                if not pd.isna(first_cell):
                    pass 
            except:
                pass
            return pd.DataFrame() 
        df.columns = df.iloc[header_idx]
        df = df[header_idx+1:]
        
        cols = pd.Series([str(c).strip() for c in df.columns])
        for dup in cols[cols.duplicated()].unique(): 
            cols[cols[cols == dup].index.values.tolist()] = [dup + '.' + str(i) if i != 0 else dup for i in range(sum(cols == dup))]
        df.columns = cols
        
        col_map = {c: c.lower().replace('\n', ' ').replace('(', '').replace(')', '').strip() for c in df.columns}
        
        # --- 2. IDENTIFY COLUMNS (FLEXIBLE) ---
        date_keywords = ["date", "txn date", "transaction date", "val. date", "tran date"]
        for col, lower_col in col_map.items():
            if any(k == lower_col for k in date_keywords) or (("date" in lower_col) and ("value" not in lower_col) and ("doc" not in lower_col)):
                self.bank_date_col = "Date"
                df.rename(columns={col: "Date"}, inplace=True)
                break
        
        ref_keywords = ["chq./ref.no.", "cheque no", "ref no", "reference", "narration", "description", "tran id", "remarks", "chqno", "particulars"]
        for col, lower_col in col_map.items():
            if any(k in lower_col for k in ref_keywords):
                self.bank_ref_col = col
                self.orig_bank_ref_col = col 
                break

        # --- 3. IDENTIFY AMOUNT & TYPE ---
        w_col = None; d_col = None; amt_col = None; type_col = None
        
        for col, lower_col in col_map.items():
            if "withdrawal" in lower_col or "debit" in lower_col or "dr amount" in lower_col: w_col = col
            if "deposit" in lower_col or "credit" in lower_col or "cr amount" in lower_col: d_col = col
            if "amount" in lower_col and "dr" not in lower_col and "cr" not in lower_col: amt_col = col
            if lower_col in ["dr/cr", "cr/dr", "type", "indicator", "d/c"]: type_col = col
        
        if w_col and d_col:
            df[w_col] = self.clean_currency(df[w_col])
            df[d_col] = self.clean_currency(df[d_col])
            df['Amount'] = df[w_col].fillna(0) + df[d_col].fillna(0)
            df['Txn_Type'] = np.where(df[w_col].fillna(0) > 0, 'Dr', 'Cr')
        elif amt_col and type_col:
            df['Amount'] = self.clean_currency(df[amt_col])
            df['Txn_Type'] = df[type_col].astype(str).str.upper().apply(lambda x: 'Cr' if 'CR' in x or 'C' == x else 'Dr')
        elif w_col and not d_col:
            df['Amount'] = self.clean_currency(df[w_col])
            df['Txn_Type'] = 'Dr'
        elif amt_col:
             df['Amount'] = self.clean_currency(df[amt_col])
             last_col = df.columns[-1]
             sample_series = df[last_col].astype(str).str.upper()
             if sample_series.str.contains('CR').any() or sample_series.str.contains('DR').any():
                 df['Txn_Type'] = sample_series.apply(lambda x: 'Cr' if 'CR' in x or 'C' == x else 'Dr')
             else:
                 df['Txn_Type'] = 'Dr' 
        else:
            return pd.DataFrame() 

        df = df[df['Amount'] > 0]
        if self.bank_date_col == "Date" and "Date" in df.columns:
            df["Date"] = pd.to_datetime(df["Date"], dayfirst=True, errors='coerce')
        if 'Txn_Type' not in df.columns: df['Txn_Type'] = 'Dr'

        df['Direction'] = df['Txn_Type'].map({'Dr': 'Outgoing', 'Cr': 'Incoming'})
        return df

    def preprocess_sap_kz_zr(self, df):
        """
        Enhances ZR rows with data from KZ rows using Clearing Document matching.
        Then filters to keep ONLY ZR rows.
        """
        # Identify critical columns (flexible search)
        col_map = {c: c.lower().strip() for c in df.columns}
        doc_type_col = next((c for c, l in col_map.items() if 'document type' in l), None)
        clear_doc_col = next((c for c, l in col_map.items() if 'clearing' in l and 'doc' in l), None)
        text_col = next((c for c, l in col_map.items() if 'text' in l), None)
        
        # If we can't find Document Type, we can't distinguish KZ/ZR, so return as is
        if not doc_type_col:
            return df

        # If we have Clearing Document, use it for robust mapping
        if clear_doc_col and text_col:
            # Create a map: Clearing Document -> Text (from KZ rows)
            kz_rows = df[df[doc_type_col].astype(str).str.upper() == 'KZ']
            # We take the first valid text found for a clearing doc
            kz_map = kz_rows.set_index(clear_doc_col)[text_col].to_dict()
            
            # Apply to ZR rows
            is_zr = df[doc_type_col].astype(str).str.upper() == 'ZR'
            
            # Function to apply mapping: If existing text is empty/generic, try to fetch from KZ
            def apply_map(row):
                val = row[clear_doc_col]
                if val in kz_map:
                    return kz_map[val] # Use KZ name
                return row[text_col] # Keep existing if no KZ match

            df.loc[is_zr, text_col] = df.loc[is_zr].apply(apply_map, axis=1)

        # Filter: "I am only considering the ZR from here"
        # We also keep rows that might not have a doc type just in case, but usually we filter strict.
        # Based on user request, strict ZR filter.
        df_filtered = df[df[doc_type_col].astype(str).str.upper() == 'ZR'].copy()
        
        return df_filtered

    def load_files(self, sheet_name=None):
        # 1. Load Bank
        if self.b.name.endswith('.csv'): 
            b_df = pd.read_csv(self.b)
        else: 
            b_df = pd.read_excel(self.b, header=None, sheet_name=sheet_name)
        
        self.df = self._prep(b_df)
        if self.df.empty:
            raise ValueError("Could not detect valid bank transactions.")

        # 2. Load SAP
        if self.s.name.endswith('.csv'): 
            self.df2 = pd.read_csv(self.s)
        else: 
            self.df2 = pd.read_excel(self.s)
        
        # --- PREPROCESS SAP: KZ -> ZR MAPPING ---
        self.df2 = self.preprocess_sap_kz_zr(self.df2)
        
        # Prepare SAP Columns
        col = "Amount in LC" if self.a == "BRS Account" else "Amount in Local Currency"
        if col not in self.df2.columns:
            found = False
            for c in self.df2.columns:
                if "amount" in c.lower() and "lc" in c.lower():
                    self.df2.rename(columns={c: col}, inplace=True)
                    found = True; break
            if not found: raise ValueError(f"Missing column '{col}' in SAP file.")
        
        self.df2[col] = self.clean_currency(self.df2[col])
        
        for c in self.df2.columns:
            if "date" in c.lower() and "doc" not in c.lower(): 
                self.sap_date_col = c
                self.df2[c] = pd.to_datetime(self.df2[c], dayfirst=True, errors='coerce')
                break

        for c in self.df2.columns:
            if c.lower() in ["assignment", "reference", "ref key", "text"]:
                self.sap_ref_col = c
                self.orig_sap_ref_col = c
                break
        
        self.sap_type_col = next((c for c in self.df2.columns if "debit" in c.lower() and "credit" in c.lower()), None)

        # 3. Load History (Already Cleared) File
        if self.h is not None:
            if self.h.name.endswith('.csv'):
                self.df_history = pd.read_csv(self.h)
            else:
                self.df_history = pd.read_excel(self.h)
            
            # Clean History Amount
            h_col = "Amount in LC" if self.a == "BRS Account" else "Amount in Local Currency" # Guessing same format
            # Fallback search for amount in history
            if h_col not in self.df_history.columns:
                for c in self.df_history.columns:
                     if "amount" in c.lower():
                         h_col = c; break
            
            if h_col in self.df_history.columns:
                self.df_history["_Amount"] = self.clean_currency(self.df_history[h_col])
            else:
                self.df_history = None # Invalid history file

    def match(self):
        bank = self.df.copy().reset_index(drop=True)
        sap = self.df2.copy().reset_index(drop=True)
        col_amt = "Amount in LC" if self.a == "BRS Account" else "Amount in Local Currency"
        
        sap["status"] = "Not Found in Bank Statement"
        sap["Match_Method"] = ""
        bank["is_matched"] = False
        sap["Direction"] = "" 
        
        if self.orig_bank_ref_col and self.orig_bank_ref_col in bank.columns:
            sap[f"Bank_{self.orig_bank_ref_col}"] = np.nan 

        if self.sap_ref_col and self.bank_ref_col:
            sap["_clean_ref"] = self.clean_ref(sap[self.sap_ref_col])
            bank["_clean_ref"] = self.clean_ref(bank[self.bank_ref_col])

        # --- MATCHING LOOP (SAP vs Bank) ---
        for i, row in sap.iterrows():
            amt = row[col_amt]
            if pd.isna(amt): continue
            
            target_type = None
            if self.sap_type_col:
                sap_ind = str(row[self.sap_type_col]).strip().upper()
                if sap_ind == 'H': 
                    target_type = 'Dr' # Outgoing
                    sap.at[i, "Direction"] = "Outgoing"
                elif sap_ind == 'S': 
                    target_type = 'Cr' # Incoming
                    sap.at[i, "Direction"] = "Incoming"
            
            if target_type:
                bank_candidates = bank[(bank['Txn_Type'] == target_type) & (bank["is_matched"] == False)]
            else:
                bank_candidates = bank[bank["is_matched"] == False]
            
            if bank_candidates.empty: continue

            # PASS 1: Ref ID
            if self.sap_ref_col and self.bank_ref_col:
                ref = row["_clean_ref"]
                if ref and len(str(ref)) > 2:
                    cand = bank_candidates[(bank_candidates["_clean_ref"]==ref) & (bank_candidates["Amount"]==amt)]
                    if not cand.empty:
                        idx = cand.index[0]
                        bank.at[idx, "is_matched"] = True
                        sap.at[i, "status"] = "Matched"
                        
                        diff_text = ""
                        if self.sap_date_col and not pd.isna(row[self.sap_date_col]):
                             diff = (bank.loc[idx, "Date"] - row[self.sap_date_col]).days
                             diff_text = f" (Diff: {diff} days)"
                        
                        sap.at[i, "Match_Method"] = f"Ref ID{diff_text}"
                        if self.orig_bank_ref_col:
                            sap.at[i, f"Bank_{self.orig_bank_ref_col}"] = bank.loc[idx, self.orig_bank_ref_col]
                        continue

            # PASS 2: Exact Date
            if self.sap_date_col and self.bank_date_col and not pd.isna(row[self.sap_date_col]):
                s_date = row[self.sap_date_col]
                cand = bank_candidates[(bank_candidates["Amount"]==amt) & (bank_candidates["Date"]==s_date)]
                if not cand.empty:
                    idx = cand.index[0]
                    bank.at[idx, "is_matched"] = True
                    sap.at[i, "status"] = "Matched"
                    sap.at[i, "Match_Method"] = "Exact Date (Diff: 0 days)"
                    if self.orig_bank_ref_col:
                        sap.at[i, f"Bank_{self.orig_bank_ref_col}"] = bank.loc[idx, self.orig_bank_ref_col]
                    continue

            # PASS 3: Soft Date
            if self.sap_date_col and self.bank_date_col and not pd.isna(row[self.sap_date_col]):
                s_date = row[self.sap_date_col]
                cand = bank_candidates[bank_candidates["Amount"]==amt]
                if not cand.empty:
                    cand = cand.copy()
                    cand["_diff"] = (cand["Date"] - s_date).dt.days.abs()
                    valid = cand[cand["_diff"] <= 3]
                    if not valid.empty:
                        idx = valid.sort_values("_diff").index[0]
                        bank.at[idx, "is_matched"] = True
                        sap.at[i, "status"] = "Matched" 
                        diff_days = (bank.loc[idx, "Date"] - s_date).days
                        sap.at[i, "Match_Method"] = f"Date Variance (Diff: {diff_days} days)"
                        if self.orig_bank_ref_col:
                            sap.at[i, f"Bank_{self.orig_bank_ref_col}"] = bank.loc[idx, self.orig_bank_ref_col]
                        continue

            # PASS 4: Amount Only
            cand = bank_candidates[bank_candidates["Amount"]==amt]
            if not cand.empty:
                idx = cand.index[0]
                bank.at[idx, "is_matched"] = True
                sap.at[i, "status"] = "Matched" 
                diff_text = "N/A"
                if self.sap_date_col and not pd.isna(row[self.sap_date_col]) and not pd.isna(bank.loc[idx, "Date"]):
                      d_val = (bank.loc[idx, "Date"] - row[self.sap_date_col]).days
                      diff_text = f"{d_val} days"
                sap.at[i, "Match_Method"] = f"Amount Only (Diff: {diff_text})"
                if self.orig_bank_ref_col:
                    sap.at[i, f"Bank_{self.orig_bank_ref_col}"] = bank.loc[idx, self.orig_bank_ref_col]

        # --- HISTORY CHECK FOR UNRECONCILED SAP ITEMS ---
        if self.df_history is not None:
             unmatched_indices = sap[sap["status"] == "Not Found in Bank Statement"].index
             
             for i in unmatched_indices:
                 amt = sap.at[i, col_amt]
                 # Look for this amount in history file
                 # Simple check: Amount match. (Could be stricter with Ref if available)
                 hist_cand = self.df_history[self.df_history["_Amount"] == amt]
                 
                 if not hist_cand.empty:
                     # Mark as Already Cleared
                     sap.at[i, "status"] = "Already Cleared"
                     sap.at[i, "Match_Method"] = "Found in History/Cleared File"

        # Handle Leftovers (Bank Only)
        unmatched = bank[bank["is_matched"]==False].copy()
        if not unmatched.empty:
            extra = pd.DataFrame({
                col_amt: unmatched["Amount"], 
                "status": "Not Found in SAP Record", 
                "Match_Method": "Bank Only (No Match)",
                "Direction": unmatched["Direction"],
            })
            target_date_col = self.sap_date_col if (self.sap_date_col and self.sap_date_col in sap.columns) else "Date"
            if self.bank_date_col and "Date" in unmatched.columns:
                 extra[target_date_col] = unmatched["Date"]
            if self.orig_sap_ref_col:
                 extra[self.orig_sap_ref_col] = np.nan 
            if self.orig_bank_ref_col: 
                 extra[f"Bank_{self.orig_bank_ref_col}"] = unmatched[self.orig_bank_ref_col]
            sap = pd.concat([sap, extra], ignore_index=True)
            
        if "_clean_ref" in sap.columns: sap.drop(columns=["_clean_ref"], inplace=True)
        self.final = sap
        
        # FINAL COLUMN REORDERING
        final_cols = list(self.final.columns)
        priority_cols = [self.sap_date_col, col_amt, "status", "Direction", "Match_Method"]
        if self.orig_sap_ref_col: priority_cols.insert(5, self.orig_sap_ref_col)
        if self.orig_bank_ref_col: priority_cols.insert(6, f"Bank_{self.orig_bank_ref_col}")

        for col in priority_cols:
            if col in final_cols: final_cols.remove(col)
        new_order = priority_cols + final_cols
        final_order = [c for c in new_order if c in self.final.columns and c is not None]
        self.final = self.final[final_order]

        # Sort and S.No.
        sort_date_col = self.sap_date_col if (self.sap_date_col and self.sap_date_col in self.final.columns) else "Date"
        if sort_date_col in self.final.columns:
            self.final[sort_date_col] = pd.to_datetime(self.final[sort_date_col], errors='coerce')
            self.final.sort_values(by=sort_date_col, ascending=True, inplace=True, na_position='last')
        self.final.reset_index(drop=True, inplace=True)
        self.final.insert(0, "S.No.", self.final.index + 1)

    def excel(self):
        buf = BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            
            # MAIN DATA
            self.final.to_excel(writer, index=False, sheet_name="Data")
            ws = writer.sheets["Data"]
            
            green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            red = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
            yellow = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
            grey = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

            col_idx = None
            for i, cell in enumerate(ws[1], start=1):
                if str(cell.value).lower() == "status": col_idx = i; break
            
            if col_idx:
                for r in range(2, ws.max_row + 1):
                    val = str(ws.cell(r, col_idx).value or "").lower()
                    if "matched" in val: ws.cell(r, col_idx).fill = green
                    elif "bank" in val: ws.cell(r, col_idx).fill = red
                    elif "sap" in val: ws.cell(r, col_idx).fill = yellow
                    elif "already" in val: ws.cell(r, col_idx).fill = grey

            # SUMMARY
            s_ws = writer.book.create_sheet("Summary")
            s_ws["A1"] = "RECONCILIATION SUMMARY"; s_ws["A1"].font = Font(bold=True, size=14)
            s_ws["A3"] = f"Total Records: {len(self.final)}"
            
            df_matched = self.final[self.final["status"] == "Matched"]
            df_cleared = self.final[self.final["status"] == "Already Cleared"]
            df_unmatched = self.final[self.final["status"].str.contains("Not Found", na=False)]
            
            s_ws["A4"] = f"Total Matched (Bank vs SAP): {len(df_matched)}"
            s_ws["A5"] = f"Already Cleared (History): {len(df_cleared)}"
            s_ws["A6"] = f"Unreconciled: {len(df_unmatched)}"

            if not df_matched.empty: df_matched.to_excel(writer, index=False, sheet_name="Matched Records")
            if not df_cleared.empty: df_cleared.to_excel(writer, index=False, sheet_name="Already Cleared")
            if not df_unmatched.empty: df_unmatched.to_excel(writer, index=False, sheet_name="Unreconciled")

            for sheet_name in writer.sheets:
                ws_curr = writer.sheets[sheet_name]
                for col in ws_curr.columns:
                    try:
                        max_len = max(len(str(c.value)) for c in col if c.value is not None)
                        ws_curr.column_dimensions[get_column_letter(col[0].column)].width = max_len + 3
                    except: pass

        buf.seek(0)
        return buf

# ----------------------------------------------------
# 6. MAIN UI LAYOUT
# ----------------------------------------------------
col_inputs, col_results = st.columns([2, 8], gap="medium") 

selected_sheet = None
bank_file = None
sap_file = None
history_file = None

with col_inputs:
    st.markdown('<div class="upload-card"><div class="card-header">🏦 Bank Statement</div>', unsafe_allow_html=True)
    bank_file = st.file_uploader("Upload Bank File", type=["xlsx", "xls", "csv"], key="bank_up", label_visibility="collapsed")
    
    if bank_file and not bank_file.name.endswith('.csv'):
        try:
            xl_file = pd.ExcelFile(bank_file)
            st.markdown("<p style='font-size:14px; margin-bottom:5px; color:#0ea5e9;'>Select Sheet:</p>", unsafe_allow_html=True)
            selected_sheet = st.selectbox("Select Sheet", xl_file.sheet_names, label_visibility="collapsed")
        except Exception as e:
            st.error(f"Error reading sheets: {e}")
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown('<div class="upload-card"><div class="card-header">💼 SAP Statement </div>', unsafe_allow_html=True)
    sap_file = st.file_uploader("Upload SAP File (KZ/ZR)", type=["xlsx", "xls", "csv"], key="sap_up", label_visibility="collapsed")
    st.markdown("</div>", unsafe_allow_html=True)
    
    st.markdown('<div class="upload-card"><div class="card-header">📂 Already Cleared / History</div>', unsafe_allow_html=True)
    history_file = st.file_uploader("Upload History File", type=["xlsx", "xls", "csv"], key="hist_up", label_visibility="collapsed")
    st.markdown("</div>", unsafe_allow_html=True)

with col_results:
    st.markdown("### 🚀 Reconciliation Dashboard")
    
    if st.session_state.processed_data is None:
        if st.button("Start Matching Process", use_container_width=True):
            if not sap_file or not bank_file:
                st.warning("⚠️ Please upload both Bank and SAP files on the left.")
            else:
                try:
                    with st.spinner("🔄 Analyzing and Matching Data..."):
                        bank_file.seek(0); sap_file.seek(0)
                        if history_file: history_file.seek(0)
                        
                        p = Processor(bank_file, sap_file, history_file, acct_type)
                        p.load_files(selected_sheet)
                        p.match()
                        
                        st.session_state.processed_data = p.final
                        st.session_state.excel_buffer = p.excel()
                        
                        st.session_state.metrics = {
                            "matched": (p.final["status"] == "Matched").sum(),
                            "cleared": (p.final["status"] == "Already Cleared").sum(),
                            "not_bank": (p.final["status"].str.contains("Not Found in Bank")).sum(),
                            "not_sap": (p.final["status"].str.contains("Not Found in SAP")).sum()
                        }
                        st.rerun()
                except Exception as e:
                    st.error(f"⚠️ Error: {str(e)}")
                    st.info("Tip: Ensure you selected the correct Sheet (not the Summary sheet).")
        
        st.markdown("""
        <div class="empty-state">
            <h3>👋 Ready to Reconcile?</h3>
            <p>Upload your files (Bank, SAP, and optional History) on the left, select the correct sheet, and click "Start Matching Process".</p>
        </div>
        """, unsafe_allow_html=True)

    else:
        m = st.session_state.metrics
        c1, c2, c3, c4 = st.columns(4)
        
        c1.metric("✅ Total Matched", m["matched"], help="Includes Exact, Date Variance, and Amount Only matches.")
        c2.metric("📁 Already Cleared", m["cleared"], help="Found in the History/Previously Cleared file.")
        c3.metric("🟥 Not in Bank", m["not_bank"], help="SAP entry exists but no matching Bank entry found.")
        c4.metric("🟨 Not in SAP", m["not_sap"], help="Bank entry exists but no matching SAP entry found.")
        
        st.divider()
        st.markdown("#### 📊 Data Preview")
        st.dataframe(st.session_state.processed_data.head(50), height=450, use_container_width=True)
        
        st.download_button(
            label="⬇️ Download Final Reconciliation Report",
            data=st.session_state.excel_buffer,
            file_name=f"Reconciliation_Report_{acct_type.replace(' ','_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
        st.markdown("""
        <div class="legend-box">
        <b>Legend:</b> 🟩 Matched &nbsp;&nbsp; ⬜ Already Cleared &nbsp;&nbsp; 🟥 Not in Bank &nbsp;&nbsp; 🟨 Not in SAP <br>
        <i>Note: Match details (days difference) are in the 'Match_Method' column.</i>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔄 Reset & Start Over", type="primary", key="reset_main"):
            reset_app()
            st.rerun()
